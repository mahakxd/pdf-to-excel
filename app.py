import gradio as gr
import pdfplumber
import pandas as pd
import google.genai as genai
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

genai_client = genai.Client(api_key="api key")

os.makedirs("output", exist_ok=True)

def extract_text(pdf_path):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                text += t + "\n"
    return text

def process_with_gemini(text):
    prompt = f"""
Convert this unstructured PDF text into a JSON array of objects (key, value, comments).
Return ONLY valid JSON.

{text}
"""
    response = genai_client.models.generate_content(
        model="gemini-2.5-flash",
        contents=prompt
    )
    raw = response.text.strip()

    if raw.startswith("```"):
        raw = raw.strip("```").strip()
    if "```json" in raw:
        raw = raw.split("```json")[-1]
    if "```" in raw:
        raw = raw.split("```")[0]

    start = raw.find("[")
    end = raw.rfind("]") + 1
    json_str = raw[start:end]

    return json.loads(json_str)

def style_excel(path):
    wb = load_workbook(path)
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    thin = Side(border_style="thin", color="000000")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for col in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = length + 5

    wb.save(path)

def process_pdf(pdf_file):
    if pdf_file is None:
        return "Upload a PDF", None, None

    text = extract_text(pdf_file)
    data = process_with_gemini(text)

    df = pd.DataFrame(data)

    excel_path = os.path.join("output", "structured_output.xlsx")
    csv_path = os.path.join("output", "structured_output.csv")

    df.to_excel(excel_path, index=False)
    style_excel(excel_path)

    df.to_csv(csv_path, index=False)

    return "Success", df, excel_path

interface = gr.Interface(
    fn=process_pdf,
    inputs=gr.File(label="Upload PDF"),
    outputs=[
        gr.Textbox(label="Status"),
        gr.Dataframe(label="CSV Preview (On-Screen)"),
        gr.File(label="Download Excel")
    ],
    title="AI PDF → Excel/CSV Document Structurer",
    description="Refined Output • Styled Excel • CSV Preview • Gemini 2.5 Flash Powered",
)

if __name__ == "__main__":
    interface.launch()
